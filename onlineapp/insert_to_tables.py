import os
import django
os.environ.setdefault("DJANGO_SETTINGS_MODULE","onlineproject.settings")
django.setup()
import click
import openpyxl
from onlineapp.models import *
def import_college_data(source,sheet_name):
    pass
    wb = openpyxl.load_workbook(source)
    sheet = wb[sheet_name]
    for row in list(sheet.rows)[1:]:
        l = []
        for cell in row:
            st = cell.value
            if st:
                l.append(st)
        if len(l):
            c = college(name=l[0],acronym=l[1],location=l[2],contact=l[3])
            c.save()
            #print(st,end=",")
        #print()
def import_Student_data(source,sheet_name):
    pass
    wb = openpyxl.load_workbook(source)
    sheet = wb[sheet_name]
    for row in list(sheet.rows)[1:]:
        l = []
        for cell in row:
            st = cell.value
            if st:
                l.append(st)
        if len(l):
            col=college.objects.get(acronym=l[1])
            c = Student(name=l[0], college=col,email=l[2],db_folder=str(l[3]).lower())
            c.save()
def import_mockest_data(source,sheet_name):
    pass
    wb = openpyxl.load_workbook(source)
    sheet = wb[sheet_name]
    for row in list(sheet.rows)[1:]:
        l = []
        for cell in row:
            st = cell.value
            if st:
                l.append(st)
        if len(l):
            pass
            stripped_db_name = l[0].split('_')[2]
            #print(stripped_db_name)
            try:
                col = Student.objects.get(db_folder=stripped_db_name)
                c = MockTest1(students=col,problem1=l[1],problem2=l[2],problem3=l[3],problem4=l[4],total=l[5])
                c.save()
            except Exception as e:
                print(e)
            #print(st,end=",")
        #print()
def import_Student_del_data(source,sheet_name):
    pass
    wb = openpyxl.load_workbook(source)
    sheet = wb[sheet_name]
    for row in list(sheet.rows)[1:]:
        l = []
        for cell in row:
            st = cell.value
            if st:
                l.append(st)
        if len(l):
            col = college.objects.get(acronym=l[1])
            c = Student(name=l[0], college=col, email=l[2], db_folder=str(l[3]).lower(),dropped_out=True)
            c.save()
'''
college_sheet = wb['Current']
for rows in range(2,college_sheet.max_row+1):
    student = []
    for cols in range(1,college_sheet.max_column+1):
        student.append(college_sheet.cell(row=rows, column=cols).value)

    try:
        col = College.objects
    except Exception as e:
        print(e)
    s = Student(name = student[0],email = student[2],db_folder = str.lower(student[3]),college =College.objects.get(acronym=student[1]))
    s.save()
'''

def display(L):
    pass
    for row in list(L):
        pass
        print(row)

def query_simple():
    pass
    L = college.objects.values_list('acronym')
    for acr in list(L):
        pass
        #print(str(acr).upper())

    queryset1 = college.objects.all()
    queryset2 = L
    #print(queryset1.query)
    #print(queryset2.query)

    count_of_colleges = college.objects.filter(location="Hyderabad").count()
    #print(count_of_colleges)

    L = college.objects.order_by('-acronym')
    for acr in list(L):
        pass
        #print(str(acr).upper())

    L = college.objects.values_list('acronym').order_by("-location")[:5]
    for acr in list(L):
        pass
        #print(str(acr).upper())

    from django.db.models import Count
    L = college.objects.values('location').annotate(Count('location'))
    for row in list(L):
        pass
        #print(row)
        #print(row['location'],"-->",row['location__count'])

def query_complex():
    pass
    L = Student.objects.filter(college__acronym="anits").values("name")
    #display(L)

    from django.db.models import Count,Avg,Max,Min

    L = Student.objects.values("college__name").annotate(count = Count("name"))
    display(L)

    L = college.objects.values('acronym').annotate(count=Count('student__id')).order_by('count')
    #display(L)

    L = college.objects.values_list('name','acronym').annotate(count=Count('student__id')).order_by('count')
    #display(L)

    L = Student.objects.values_list('name','email','college__name')
    #display(L)

    L = college.objects.values('acronym').annotate(count=Count('student__id')).filter(count__gte=10).order_by('-count')
    #display(L)

    L = MockTest1.objects.values('total','students__name','students__college__name')
    #display(L)

    L = Student.objects.values('college__name','mocktest1__total')##,'__total')#.annotate(avg=Avg('students__total')).order_by('avg')
    #display(L)

def import_data():
    source = 'students.xlsx'
    source_sheet = 'Colleges'
    import_college_data(source,source_sheet)
    source = 'students.xlsx'
    source_sheet = 'Current'
    import_Student_data(source,source_sheet)
    source = 'html_data.xlsx'
    source_sheet = 'mocktest_details'
    import_mockest_data(source,source_sheet)
    source = 'students.xlsx'
    source_sheet = 'Deletions'
    import_Student_del_data(source, source_sheet)
#import_data()
#query_simple()
query_complex()